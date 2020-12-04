import openpyxl
wb = openpyxl.load_workbook('E.xlsx')
# wb = Workbook()
# mySheet = wb.get_sheet_names()[0]
# sheetDate = mySheet.split()[2]
# sheet_1 = wb.get_sheet_names()
sheet_1 = wb.sheetnames

# print(sheet_1)
mySheet = wb[sheet_1[0]]
# print(mySheet.max_row, mySheet.max_column)
myCell = mySheet.cell(row=5, column=1)
print(myCell)